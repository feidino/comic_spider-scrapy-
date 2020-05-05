# # 引用requests库   
# import requests
# # 调用get方法，下载这个字典
# res_music = requests.get('https://c.y.qq.com/soso/fcgi-bin/client_search_cp?ct=24&qqmusic_ver=1298&new_json=1&remoteplace=txt.yqq.song&searchid=60997426243444153&t=0&aggr=1&cr=1&catZhida=1&lossless=0&flag_qc=0&p=1&n=20&w=%E5%91%A8%E6%9D%B0%E4%BC%A6&g_tk=5381&loginUin=0&hostUin=0&format=json&inCharset=utf8&outCharset=utf-8&notice=0&platform=yqq.json&needNewCode=0')
# # 使用json()方法，将response对象，转为列表/字典
# json_music = res_music.json()
# # 一层一层地取字典，获取歌单列表
# list_music = json_music['data']['song']['list']
# # list_music是一个列表，music是它里面的元素
# for music in list_music:
#     # 以name为键，查找歌曲名
#     print(music['name'])
#     # 查找专辑名
#     print('所属专辑：'+music['album']['name'])
#     # 查找播放时长
#     print('播放时长：'+str(music['interval'])+'秒')
#     # 查找播放链接
#     print('播放链接：https://y.qq.com/n/yqq/song/'+music['mid']+'.html\n\n')


import requests
# # 引用requests模块
# for i in range(5):
#     res_comments = requests.get('''https://c.y.qq.com/base/fcgi-bin/fcg_global_comment_h5.fcg?g_tk=5381&loginUin=0&hostUin=0
#     &format=json&inCharset=utf8&outCharset=GB2312&notice=0&platform=yqq.json&needNewCode=0&cid=205360772&reqtype=2&biztype=1&topid
#     =102065756&cmd=6&needmusiccrit=0&pagenum=%d&pagesize=15&lasthotcommentid=song_102065756_3202544866_44059185&domain=qq.com&ct=24&
#     cv=10101010'''%i)
#     # 调用get方法，下载评论列表
#     json_comments = res_comments.json()
#     # 使用json()方法，将response对象，转为列表/字典
#     list_comments = json_comments['comment']['commentlist']
#     # 一层一层地取字典，获取评论列表
#     for comment in list_comments:
#     # list_comments是一个列表，comment是它里面的元素
#         print(comment['rootcommentcontent'])
#         # 输出评论
#         print('-----------------------------------')
#         # 将不同的评论分隔开来


# for i in range(5):
#     res_comments = requests.get('https://c.y.qq.com/base/fcgi-bin/fcg_global_comment_h5.fcg?g_tk=5381&loginUin=0&hostUin=0&format=json&inCharset=utf8&outCharset=GB2312&notice=0&platform=yqq.json&needNewCode=0&cid=205360772&reqtype=2&biztype=1&topid=102065756&cmd=6&needmusiccrit=0&pagenum='+str(i)+'&pagesize=15&lasthotcommentid=song_102065756_3202544866_44059185&domain=qq.com&ct=24&cv=10101010')
#     # 调用get方法，下载评论列表
#     json_comments = res_comments.json()
#     # 使用json()方法，将response对象，转为列表/字典
#     list_comments = json_comments['comment']['commentlist']
#     # 一层一层地取字典，获取评论列表
#     for comment in list_comments:
#     # list_comments是一个列表，comment是它里面的元素
#         print(comment['rootcommentcontent'])
#         # 输出评论
#         print('-----------------------------------')
#         # 将不同的评论分隔开来
# import openpyxl 
# wb=openpyxl.Workbook('Marvel.xlsx') 
# sheet=wb.active
# # sheet.title='new title'
# sheet['A1'] = 'footballclub'
# rows= [['利物浦','曼娘','阿森纳'],['红军','红魔','兵工厂']]
# for i in rows:
#     sheet.append(i)
# print(rows)
import openpyxl 
wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb.create_sheet('german')
sheet['A1'] = '多特蒙德'
sheet['A2'] = '拜仁慕尼黑'
# rows= [['利物浦','曼娘','阿森纳'],['红军','红魔','兵工厂']]
# for i in rows:
#     sheet.append(i)
# print(rows)
wb.save('Marvel.xlsx')

# print(sheetname)
# A1_cell = sheet['A1']
# A1_value = A1_cell.value
# print(A1_value)

# import csv
# csv_file = open('demo.csv','w',newline='',encoding='gbk')
# csv_write = csv.writer(csv_file)
# csv_write.writerow(['利物浦','曼联','阿森纳','车尔西','曼城','热刺'])
# csv_file.close()

# import requests,openpyxl 
# wb=openpyxl.Workbook() 
# sheet=wb.active
# sheet.title='jay_song'
# rows= ['歌曲名称','所属专辑','播放时长','播放链接']
# sheet.append(rows)
# url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
# for x in range(5):

#     params = {
#         'ct': '24',
#         'qqmusic_ver': '1298',
#         'new_json': '1',
#         'remoteplace': 'sizer.yqq.song_next',
#         'searchid': '64405487069162918',
#         't': '0',
#         'aggr': '1',
#         'cr': '1',
#         'catZhida': '1',
#         'lossless': '0',
#         'flag_qc': '0',
#         'p': str(x + 1),
#         'n': '20',
#         'w': '周杰伦',
#         'g_tk': '5381',
#         'loginUin': '0',
#         'hostUin': '0',
#         'format': 'json',
#         'inCharset': 'utf8',
#         'outCharset': 'utf-8',
#         'notice': '0',
#         'platform': 'yqq.json',
#         'needNewCode': '0'
#     }

#     res_music = requests.get(url, params=params)
#     json_music = res_music.json()
#     list_music = json_music['data']['song']['list']
#     song_info = []
#     for music in list_music:
#         musi_tim = str(music['interval']) + '秒'
#         musi_link = 'https://y.qq.com/n/yqq/song/' + music['file']['media_mid'] + '.html'
#         song_info = [music['name'],music['album']['name'],musi_tim,musi_link]
#         sheet.append(song_info)
#         print(music['name'])
#         print('所属专辑：' + music['album']['name'])
#         print('播放时长：' + str(music['interval']) + '秒')
#         print('播放链接：https://y.qq.com/n/yqq/song/' + music['file']['media_mid'] + '.html\n\n')